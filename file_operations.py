import pdfplumber
import openpyxl


def open_first_pdf(selected_pdf):
    """ Function for specific file. Only works properly if format of the file is not changed.

    :param selected_pdf: path to PDF to be parsed
    :return:
    """
    pdf = pdfplumber.open(selected_pdf)
    page = pdf.pages[0]
    table = page.extract_table()
    new_table = table[2:]
    formated_table = ["" * 17 for i in range(len(new_table))]
    positions = [0, 1, 13, 8, 7]
    for i in range(len(new_table)):
        new_table[i][3] = new_table[i][3].replace(',', '.')
        new_table[i][4] = new_table[i][4].replace(',', '.')

    return new_table


def write_final_excel(file_path, save_path, broj_vagona=None, otpremna_zelj_uprava=None, sifra_otpremnog_kol=None,
                      uputna_zelj_uprava=None, sifra_uputnog_kol=None, okvirni_opis_tereta=None,
                      duzina_vagona=None, tara_vagona=None, neto_vagona=None, rucno_kocena_tezina=None,
                      zracno_kocena_tezina=None, slovna_serija=None, broj_osovina=None, otpremni_kolodvor=None,
                      uputni_kolodvor=None):
    """ Function for writing final XLSM file.

    :param file_path:
    :param save_path:
    :param broj_vagona:
    :param otpremna_zelj_uprava:
    :param sifra_otpremnog_kol:
    :param uputna_zelj_uprava:
    :param sifra_uputnog_kol:
    :param okvirni_opis_tereta:
    :param duzina_vagona:
    :param tara_vagona:
    :param neto_vagona:
    :param rucno_kocena_tezina:
    :param zracno_kocena_tezina:
    :param slovna_serija:
    :param broj_osovina:
    :param otpremni_kolodvor:
    :param uputni_kolodvor:
    :return:
    """
    workbook = openpyxl.load_workbook(file_path, keep_vba=True)
    worksheet = workbook["Sheet1"]

    if broj_vagona is not None:
        for i in range(len(broj_vagona)):
            worksheet.cell(row=10+i, column=8, value=broj_vagona[i])

    if otpremna_zelj_uprava is not None:
        for i in range(len(otpremna_zelj_uprava)):
            worksheet.cell(row=10+i, column=10, value=otpremna_zelj_uprava[i])

    if sifra_otpremnog_kol is not None:
        for i in range(len(sifra_otpremnog_kol)):
            worksheet.cell(row=10+i, column=12, value=sifra_otpremnog_kol[i])

    if uputna_zelj_uprava is not None:
        for i in range(len(uputna_zelj_uprava)):
            worksheet.cell(row=10+i, column=14, value=uputna_zelj_uprava[i])

    if sifra_uputnog_kol is not None:
        for i in range(len(sifra_uputnog_kol)):
            worksheet.cell(row=10 + i, column=16, value=sifra_uputnog_kol[i])

    if okvirni_opis_tereta is not None:
        for i in range(len(okvirni_opis_tereta)):
            worksheet.cell(row=10 + i, column=22, value=okvirni_opis_tereta[i])

    if duzina_vagona is not None:
        for i in range(len(duzina_vagona)):
            worksheet.cell(row=10 + i, column=24, value=duzina_vagona[i])

    if tara_vagona is not None:
        for i in range(len(tara_vagona)):
            worksheet.cell(row=10 + i, column=26, value=tara_vagona[i])

    if neto_vagona is not None:
        for i in range(len(neto_vagona)):
            worksheet.cell(row=10 + i, column=28, value=neto_vagona[i])

    if rucno_kocena_tezina is not None:
        for i in range(len(rucno_kocena_tezina)):
            worksheet.cell(row=10 + i, column=30, value=rucno_kocena_tezina[i])

    if zracno_kocena_tezina is not None:
        for i in range(len(zracno_kocena_tezina)):
            worksheet.cell(row=10 + i, column=34, value=zracno_kocena_tezina[i])

    if slovna_serija is not None:
        for i in range(len(slovna_serija)):
            worksheet.cell(row=10 + i, column=38, value=slovna_serija[i])

    if broj_osovina is not None:
        for i in range(len(broj_osovina)):
            worksheet.cell(row=10 + i, column=40, value=broj_osovina[i])

    if otpremni_kolodvor is not None:
        for i in range(len(otpremni_kolodvor)):
            worksheet.cell(row=10 + i, column=42, value=otpremni_kolodvor[i])

    if uputni_kolodvor is not None:
        for i in range(len(uputni_kolodvor)):
            worksheet.cell(row=10 + i, column=44, value=uputni_kolodvor[i])

    workbook.save(save_path)